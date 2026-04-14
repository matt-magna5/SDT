"""
parse_ntnx_collector.py — Nutanix Collector XLS → vsphere-perf JSON
Reads a ntnxcollector_*.xlsx file and outputs a standardized perf JSON
compatible with collect_vsphere_perf.py output format.

Usage: python parse_ntnx_collector.py <path_to_xlsx> [output_dir]
If output_dir omitted, writes next to the XLS file.
"""
import openpyxl, json, sys, os, re
from datetime import datetime

def sheet_rows(wb, name):
    """Return list of dicts for a sheet."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

def safe(v, default=None):
    return v if v is not None else default

def run(xlsx_path, output_dir=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Metadata ──────────────────────────────────────────────────────────────
    meta_rows = sheet_rows(wb, 'Metadata')
    meta = meta_rows[0] if meta_rows else {}
    collected_at = str(meta.get('Collection Date & Time', datetime.now().isoformat()))
    duration_days_str = str(meta.get('Performance Data Duration', '90 Days'))
    duration_days = int(re.search(r'\d+', duration_days_str).group()) if re.search(r'\d+', duration_days_str) else 90
    source_label = f"Nutanix Collector v{meta.get('Collector Version','?')} — {duration_days} days"

    # ── Host ──────────────────────────────────────────────────────────────────
    host_rows = sheet_rows(wb, 'vHosts')
    h = host_rows[0] if host_rows else {}
    host = {
        "Name":           safe(h.get('Host Name'), ''),
        "Model":          safe(h.get('Model'), ''),
        "ServiceTag":     safe(h.get('Service Tag'), ''),
        "Cores":          safe(h.get('CPU Cores'), 0),
        "CPUModel":       safe(h.get('CPU Model'), ''),
        "CPUSpeedMHz":    safe(h.get('CPU Speed'), 0),
        "RAMgb":          safe(h.get('Memory Size'), 0),
        "Hypervisor":     safe(h.get('Hypervisor'), ''),
        "NICs":           safe(h.get('NICs'), 0),
        "CPUUsagePct":    safe(h.get('CPU Usage'), 0),
        "MemUsagePct":    safe(h.get('Memory Usage'), 0),
        "IOPS_95th":      safe(h.get('95th Percentile IOPS'), 0),
        "DiskKBps_95th":  safe(h.get('95th Percentile Disk Throughput (KBps)'), 0),
        "VMCount":        safe(h.get('VMs'), 0),
    }

    # ── Cluster ───────────────────────────────────────────────────────────────
    cluster_rows = sheet_rows(wb, 'vCluster')
    c = cluster_rows[0] if cluster_rows else {}
    cluster = {
        "Name":          safe(c.get('Cluster Name'), host["Name"]),
        "CPUUsagePct":   safe(c.get('CPU Usage %'), 0),
        "MemUsagePct":   safe(c.get('Memory Usage %'), 0),
        "IOPS_95th":     safe(c.get('95th Percentile IOPS'), 0),
        "DiskKBps_95th": safe(c.get('95th Percentile Disk Throughput (KBps)'), 0),
        "CapacityMiB":   safe(c.get('Capacity (MiB)'), 0),
        "ConsumedMiB":   safe(c.get('Consumed (MiB)'), 0),
    }

    # ── Per-VM data ───────────────────────────────────────────────────────────
    vcpu_map  = {r['VM Name']: r for r in sheet_rows(wb, 'vCPU')}
    vmem_map  = {r['VM Name']: r for r in sheet_rows(wb, 'vMemory')}
    vinfo_map = {r['VM Name']: r for r in sheet_rows(wb, 'vInfo')}
    vmlist    = {r['VM Name']: r for r in sheet_rows(wb, 'vmList')}
    vpart_map = {}
    for row in sheet_rows(wb, 'vPartition'):
        vname = row['VM Name']
        if vname not in vpart_map:
            vpart_map[vname] = []
        vpart_map[vname].append(row)
    vsw_map   = {r['VM Name']: r for r in sheet_rows(wb, 'vNetwork')}

    vms = []
    for name, cpu in vcpu_map.items():
        mem  = vmem_map.get(name, {})
        info = vinfo_map.get(name, {})
        vml  = vmlist.get(name, {})
        parts = vpart_map.get(name, [])

        # OS detection — prefer VMware Tools guest name
        guest_os = safe(info.get('Guest OS'), '')
        is_linux = bool(guest_os and any(k in guest_os.lower() for k in
                   ['linux','ubuntu','centos','rhel','debian','photon','suse','oracle',
                    'rocky','alma','amazon','coreos','fedora']))
        if not is_linux and parts:
            # fallback: Linux mounts have '/' paths
            is_linux = any(p.get('Path','').startswith('/') for p in parts if p.get('Path'))

        ram_mib = safe(mem.get('Size (MiB)'), 0) or 0
        cap_mib = safe(vml.get('Capacity (MiB)'), 0) or 0
        cons_mib = safe(vml.get('Consumed (MiB)'), 0) or 0

        # Partition detail
        partition_detail = []
        for p in parts:
            partition_detail.append({
                "Path":        safe(p.get('Path'), ''),
                "CapacityMiB": safe(p.get('Capacity (MiB)'), 0),
                "ConsumedMiB": safe(p.get('Consumed (MiB)'), 0),
            })

        vm = {
            "Name":        name,
            "MOID":        safe(info.get('MOID'), ''),
            "PowerState":  safe(cpu.get('Power State'), safe(info.get('Power State'), '')),
            "GuestOS":     guest_os,
            "IsLinux":     is_linux,
            "vCPUs":       safe(cpu.get('vCPUs'), 0),
            "RAMgb":       round(ram_mib / 1024, 1) if ram_mib else 0,
            "DiskCapGB":   round(cap_mib / 1024, 1) if cap_mib else 0,
            "DiskConsumedGB": round(cons_mib / 1024, 1) if cons_mib else 0,
            "ToolStatus":  safe(info.get('Tool Status'), ''),
            "CPU": {
                "Average": safe(cpu.get('Average %')),
                "Peak":    safe(cpu.get('Peak %')),
                "Median":  safe(cpu.get('Median %')),
                "P95":     safe(cpu.get('95th Percentile % (recommended)')),
                "Ready_P95": safe(cpu.get('CPU Readiness 95th Percentile %')),
            },
            "Memory": {
                "Average": safe(mem.get('Average %')),
                "Peak":    safe(mem.get('Peak %')),
                "Median":  safe(mem.get('Median %')),
                "P95":     safe(mem.get('95th Percentile % (recommended)')),
            },
            "IOPS_P95":    safe(info.get('95th Percentile IOPS')),
            "Partitions":  partition_detail,
            "Datastore":   safe(vml.get('Datastore'), ''),
        }
        vms.append(vm)

    # ── Datastores ────────────────────────────────────────────────────────────
    datastores = []
    for row in sheet_rows(wb, 'Datastore'):
        cap  = safe(row.get('Capacity (MiB)'), 0) or 0
        cons = safe(row.get('Consumed (MiB)'), 0) or 0
        datastores.append({
            "Name":       safe(row.get('Datastore Name'), ''),
            "Type":       safe(row.get('Datastore Type'), ''),
            "CapacityGB": round(cap / 1024, 1),
            "ConsumedGB": round(cons / 1024, 1),
            "FreeGB":     round((cap - cons) / 1024, 1),
            "UsedPct":    round(cons / cap * 100, 1) if cap else 0,
        })

    # ── Licenses ──────────────────────────────────────────────────────────────
    licenses = []
    for row in sheet_rows(wb, 'vLicense'):
        licenses.append({
            "Name":    safe(row.get('Name'), ''),
            "Used":    safe(row.get('Used'), 0),
            "Total":   str(safe(row.get('Total'), '')),
            "Expiry":  str(safe(row.get('Expiry Date'), '')),
            "Key":     safe(row.get('License Key'), ''),
        })

    # ── Assemble output ───────────────────────────────────────────────────────
    out = {
        "_type":        "vSpherePerf",
        "_source":      source_label,
        "CollectedAt":  collected_at,
        "DurationDays": duration_days,
        "Host":         host,
        "Cluster":      cluster,
        "VMs":          vms,
        "Datastores":   datastores,
        "Licenses":     licenses,
    }

    # ── Write output ──────────────────────────────────────────────────────────
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsx_path))
    date_str = datetime.now().strftime('%Y-%m-%d')
    out_path = os.path.join(output_dir, f'vsphere-perf-{date_str}.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, default=str)
    print(f"Written: {out_path}")
    return out_path

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python parse_ntnx_collector.py <xlsx_path> [output_dir]")
        sys.exit(1)
    xlsx = sys.argv[1]
    odir = sys.argv[2] if len(sys.argv) > 2 else None
    run(xlsx, odir)
